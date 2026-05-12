# ─────────────────────────────────────────────────────────────────────────────
# analyze_who_eats_whom.py
#
# Purpose:
#   Analyzes the cleaned pipe-delimited log file for the "Who Eats Whom"
#   food web game. Produces per-session and overall stats.
#
#   Session splitting strategy:
#     1. Split rows by IP
#     2. Within each IP, split further by device (browser type)
#     3. Within each IP+device stream, split by TUTORIAL_STARTED events
#        (also supports legacy STARTED events from before dual-timestamp logging)
#     4. All sessions sorted by start time at the end
#
#   Two timestamps are tracked per session:
#     - TUTORIAL_STARTED: page load / tutorial shown (person walks up)
#     - GAME_STARTED:     tutorial dismissed (person starts playing)
#   This allows separating idle tutorial dwell time from actual gameplay.
#   All timestamps are converted from UTC to EDT (UTC-4).
#
# Usage:
#   python3 analyze_who_eats_whom.py <cleaned_game_file>
#   Example: python3 analyze_who_eats_whom.py cleaned/who-eats-whom.csv
#
# Output:
#   Prints analysis to console and saves a summary Excel file in cleaned/
#
# ─────────────────────────────────────────────────────────────────────────────
# ASSUMPTIONS & ANALYSIS DECISIONS
# ─────────────────────────────────────────────────────────────────────────────
#
# 1. SESSION IDENTIFICATION
#    A new session is defined by a TUTORIAL_STARTED log event (or legacy STARTED).
#    This fires every time the game page loads.
#
# 2. DUAL TIMESTAMPS
#    TUTORIAL_STARTED = page load. The tutorial overlay appears and may sit idle
#    for a long time at exhibitions between visitors.
#    GAME_STARTED = fired when the user dismisses the tutorial (skip or finish).
#    Sessions where GAME_STARTED is missing means the visitor left before
#    touching the tutorial dismiss button.
#    tutorial_duration = GAME_STARTED - TUTORIAL_STARTED (tutorial idle time)
#    game_duration     = last event   - GAME_STARTED      (actual play time)
#    total_duration    = last event   - TUTORIAL_STARTED   (full session)
#
# 3. DEVICE SEPARATION (shared WiFi)
#    The event was hosted on a shared WiFi network, meaning multiple devices
#    could share the same IP address. Logs from different devices were therefore
#    interleaved under the same IP in some sessions.
#    To handle this, sessions are first split by IP and then further split by
#    device type extracted from the browser user-agent string (Windows / Mac /
#    iPad / iPhone / Android). Each unique IP+device combination is treated as
#    a separate independent stream before session splitting is applied.
#
# 4. CASCADE GROUPING
#    DELETED_CASCADE events that occur within CASCADE_GAP_SECONDS of each
#    other are treated as a single cascade event (one removal triggering a
#    chain). Events separated by more than this threshold are counted as
#    distinct cascades. CASCADE_GAP_SECONDS is set to 10s.
#
# 5. TIME TO SUN
#    Measured from GAME_STARTED (or TUTORIAL_STARTED if no GAME_STARTED) to
#    the first ADDED event for "Sun". Sessions where Sun was never added: None.
#
# 6. FIRST ANIMAL AFTER SUN
#    The first ADDED event (for any animal other than "Sun") that occurs after
#    the first Sun ADDED event in the session.
#
# 7. TIME TO CASCADE
#    Measured from GAME_STARTED (or TUTORIAL_STARTED if no GAME_STARTED) to
#    the first DELETED_CASCADE event. Sessions with no cascades: None.
#
# 8. TIMEZONE
#    All timestamps in the raw logs are in UTC (from Loki/Grafana).
#    All times in this analysis are converted to EDT (UTC-4), which was the
#    local time at the Marbles Kids Museum event on April 11, 2026 in Raleigh NC.
#
# 9. TRUNCATED LOG FILE
#    The raw log export from Grafana was limited to 1000 rows. Browser strings
#    truncated at "KHTML" do not affect any analysis fields.
# ─────────────────────────────────────────────────────────────────────────────

import csv
import os
import sys
from datetime import datetime, timezone, timedelta
from collections import defaultdict, Counter

CASCADE_GAP_SECONDS = 10  # max gap between DELETED_CASCADE events in one cascade

if len(sys.argv) != 2:
    print("Usage: python3 analyze_who_eats_whom.py <cleaned_game_file>")
    sys.exit(1)

input_file = sys.argv[1]
output_dir = os.path.dirname(input_file)

EDT = timezone(timedelta(hours=-4))


def get_device(browser: str) -> str:
    if "iPad" in browser:      return "iPad"
    if "iPhone" in browser:    return "iPhone"
    if "Android" in browser:   return "Android"
    if "Windows" in browser:   return "Windows"
    if "Macintosh" in browser: return "Mac"
    return "Unknown"


def group_cascades(cascade_events: list) -> list[list]:
    """Group consecutive DELETED_CASCADE events into individual cascade bursts."""
    if not cascade_events:
        return []
    groups = []
    current_group = [cascade_events[0]]
    for ev in cascade_events[1:]:
        gap = (ev["timestamp"] - current_group[-1]["timestamp"]).total_seconds()
        if gap <= CASCADE_GAP_SECONDS:
            current_group.append(ev)
        else:
            groups.append(current_group)
            current_group = [ev]
    groups.append(current_group)
    return groups


# ── Load data ─────────────────────────────────────────────────────────────────
rows = []
with open(input_file, newline='', encoding='utf-8') as f:
    reader = csv.DictReader(f, delimiter="|")
    for row in reader:
        row = {k.strip(): v.strip() for k, v in row.items()}
        row["timestamp"] = datetime.fromisoformat(row["timestamp"].replace("Z", "+00:00")).astimezone(EDT)
        row["device"] = get_device(row["browser"])
        rows.append(row)

rows.sort(key=lambda r: r["timestamp"])

# ── Step 1: Split by IP, then by device ───────────────────────────────────────
streams: dict = defaultdict(list)
for row in rows:
    key = (row["ip"], row["device"])
    streams[key].append(row)

# ── Step 2: Within each stream, split by TUTORIAL_STARTED (or legacy STARTED) ─
SESSION_START_ACTIONS = {"TUTORIAL_STARTED", "STARTED"}

all_sessions = []
for (ip, device), stream_rows in streams.items():
    current = []
    for row in stream_rows:
        if row["action"] in SESSION_START_ACTIONS:
            if current:
                all_sessions.append(current)
            current = [row]
        else:
            current.append(row)
    if current:
        all_sessions.append(current)

# ── Step 3: Sort all sessions by start time ───────────────────────────────────
all_sessions.sort(key=lambda s: s[0]["timestamp"])

print(f"{'='*70}")
print(f"  WHO EATS WHOM — LOG ANALYSIS  (all times in EDT)")
print(f"{'='*70}\n")

# ── Compute per-session metrics ───────────────────────────────────────────────
session_stats = []

for i, session in enumerate(all_sessions, 1):
    tutorial_start = session[0]["timestamp"]
    ip             = session[0]["ip"]
    device         = session[0]["device"]

    # GAME_STARTED = when tutorial was dismissed
    game_started_events = [r for r in session if r["action"] == "GAME_STARTED"]
    game_start = game_started_events[0]["timestamp"] if game_started_events else None

    # Use game_start as reference for time-to-X metrics; fall back to tutorial_start
    ref_start = game_start if game_start else tutorial_start

    play_events = [r for r in session if r["action"] not in SESSION_START_ACTIONS and r["action"] != "GAME_STARTED"]
    time_out = play_events[-1]["timestamp"] if play_events else (game_start or tutorial_start)

    tutorial_duration = (game_start - tutorial_start).total_seconds() if game_start else None
    game_duration     = (time_out - game_start).total_seconds() if game_start and play_events else None
    total_duration    = (time_out - tutorial_start).total_seconds()

    adds     = [r for r in session if r["action"] == "ADDED"]
    deletes  = [r for r in session if r["action"] == "DELETED"]
    cascades_raw = [r for r in session if r["action"] == "DELETED_CASCADE"]

    # Time to first Sun added (measured from game start)
    sun_events = [r for r in adds if r["animal"] == "Sun"]
    time_to_sun = (sun_events[0]["timestamp"] - ref_start).total_seconds() if sun_events else None

    # First animal added after Sun
    first_after_sun = None
    if sun_events:
        sun_time = sun_events[0]["timestamp"]
        post_sun_adds = [r for r in adds if r["timestamp"] > sun_time and r["animal"] != "Sun"]
        first_after_sun = post_sun_adds[0]["animal"] if post_sun_adds else None

    # Animal frequency in this session
    add_counter    = Counter(r["animal"] for r in adds)
    delete_counter = Counter(r["animal"] for r in deletes + cascades_raw)

    most_deleted = delete_counter.most_common(1)[0] if delete_counter else None

    # Cascade grouping
    cascade_groups = group_cascades(cascades_raw)
    cascade_sizes  = [len(g) for g in cascade_groups]
    longest_cascade_idx = cascade_sizes.index(max(cascade_sizes)) if cascade_sizes else None

    # Time to first cascade (measured from game start)
    time_to_cascade = (cascades_raw[0]["timestamp"] - ref_start).total_seconds() if cascades_raw else None

    session_stats.append({
        "session_num":         i,
        "ip":                  ip,
        "device":              device,
        "tutorial_start":      tutorial_start,
        "game_start":          game_start,
        "time_out":            time_out,
        "tutorial_duration":   tutorial_duration,
        "game_duration":       game_duration,
        "total_duration":      total_duration,
        "adds":                adds,
        "deletes":             deletes,
        "cascades_raw":        cascades_raw,
        "cascade_groups":      cascade_groups,
        "cascade_sizes":       cascade_sizes,
        "longest_cascade_idx": longest_cascade_idx,
        "time_to_sun":         time_to_sun,
        "first_after_sun":     first_after_sun,
        "add_counter":         add_counter,
        "delete_counter":      delete_counter,
        "most_deleted":        most_deleted,
        "time_to_cascade":     time_to_cascade,
        "all_events":          session,
    })

# ── PER SESSION OUTPUT ────────────────────────────────────────────────────────
print(f"PER SESSION")
print(f"{'-'*70}")

for s in session_stats:
    i               = s["session_num"]
    tutorial_start  = s["tutorial_start"]
    game_start      = s["game_start"]
    time_out        = s["time_out"]
    device          = s["device"]
    ip              = s["ip"]

    print(f"\n  Session {i}  |  {device}  |  IP: {ip}")
    print(f"    Tutorial started: {tutorial_start.strftime('%H:%M:%S EDT')}")
    if game_start:
        tut_dur = s["tutorial_duration"]
        print(f"    Game started:     {game_start.strftime('%H:%M:%S EDT')}  ({tut_dur:.0f}s in tutorial)")
    else:
        print(f"    Game started:     — (tutorial never dismissed)")
    print(f"    Time out:         {time_out.strftime('%H:%M:%S EDT')}")
    if s["game_duration"] is not None:
        print(f"    Play time:        {s['game_duration']:.0f}s  (total incl. tutorial: {s['total_duration']:.0f}s)")
    else:
        print(f"    Total time:       {s['total_duration']:.0f}s  (no gameplay recorded)")

    # Time to Sun
    if s["time_to_sun"] is not None:
        print(f"    Time from game start to Sun added: {s['time_to_sun']:.1f}s")
    else:
        print(f"    Sun never added")

    # First animal after Sun
    if s["first_after_sun"]:
        print(f"    First animal added after Sun: {s['first_after_sun']}")
    else:
        print(f"    No animals added after Sun (or Sun never added)")

    # Event order
    print(f"    Event order:")
    for ev in s["all_events"]:
        if ev["action"] in SESSION_START_ACTIONS or ev["action"] == "GAME_STARTED":
            continue
        tag = ev["action"]
        print(f"      {ev['timestamp'].strftime('%H:%M:%S')}  {tag:<18} {ev['animal']}")

    # Animal frequency
    print(f"    Animal add counts:")
    for animal, count in s["add_counter"].most_common():
        print(f"      {count:>3}x  ADDED     {animal}")
    print(f"    Animal delete counts (DELETED + CASCADE):")
    for animal, count in s["delete_counter"].most_common():
        print(f"      {count:>3}x  DELETED   {animal}")

    # Most deleted
    if s["most_deleted"]:
        print(f"    Most deleted animal: {s['most_deleted'][0]} ({s['most_deleted'][1]}x)")

    # Cascades
    if not s["cascade_groups"]:
        print(f"    No cascades")
    else:
        print(f"    Cascades: {len(s['cascade_groups'])} total")
        for ci, group in enumerate(s["cascade_groups"], 1):
            animals = [r["animal"] for r in group]
            longest_flag = " ← LONGEST" if ci - 1 == s["longest_cascade_idx"] and len(s["cascade_groups"]) > 1 else ""
            print(f"      Cascade {ci} ({len(group)} animals lost){longest_flag}: {', '.join(animals)}")
        print(f"    Longest cascade: {max(s['cascade_sizes'])} animals")

        if s["time_to_cascade"] is not None:
            print(f"    Time to first cascade: {s['time_to_cascade']:.1f}s from session start")

    # Cascade deletions breakdown
    if s["cascades_raw"]:
        print(f"    Cascade-deleted animals (no food source):")
        for r in s["cascades_raw"]:
            print(f"      {r['timestamp'].strftime('%H:%M:%S')}  {r['animal']}")

# ── OVERALL STATS ─────────────────────────────────────────────────────────────
print(f"\n\n{'='*70}")
print(f"  OVERALL STATS")
print(f"{'='*70}\n")

total_sessions = len(session_stats)
device_counter = Counter(s["device"] for s in session_stats)

sessions_with_game_start = sum(1 for s in session_stats if s["game_start"] is not None)

tutorial_durations  = [s["tutorial_duration"] for s in session_stats if s["tutorial_duration"] is not None]
avg_tutorial_dur    = sum(tutorial_durations) / len(tutorial_durations) if tutorial_durations else 0

game_durations      = [s["game_duration"] for s in session_stats if s["game_duration"] is not None]
avg_game_dur        = sum(game_durations) / len(game_durations) if game_durations else 0

total_durations     = [s["total_duration"] for s in session_stats]
avg_total_dur       = sum(total_durations) / len(total_durations) if total_durations else 0

sun_times           = [s["time_to_sun"] for s in session_stats if s["time_to_sun"] is not None]
avg_time_to_sun     = sum(sun_times) / len(sun_times) if sun_times else 0

sessions_with_cascade = sum(1 for s in session_stats if s["cascade_groups"])
cascade_count_dist    = Counter(len(s["cascade_groups"]) for s in session_stats)

all_cascade_times   = [s["time_to_cascade"] for s in session_stats if s["time_to_cascade"] is not None]
avg_time_to_cascade = sum(all_cascade_times) / len(all_cascade_times) if all_cascade_times else 0

all_cascade_sizes   = [size for s in session_stats for size in s["cascade_sizes"]]
avg_cascade_size    = sum(all_cascade_sizes) / len(all_cascade_sizes) if all_cascade_sizes else 0

first_after_sun_counter = Counter(
    s["first_after_sun"] for s in session_stats if s["first_after_sun"]
)

overall_add_counter    = Counter()
overall_delete_counter = Counter()
for s in session_stats:
    overall_add_counter    += s["add_counter"]
    for r in s["deletes"] + s["cascades_raw"]:
        overall_delete_counter[r["animal"]] += 1

print(f"SESSIONS")
print(f"{'-'*70}")
print(f"  Total sessions:              {total_sessions}")
print(f"  Sessions by device:")
for device, count in device_counter.most_common():
    print(f"    {device:<12} {count} sessions")

print(f"\nTIME METRICS")
print(f"{'-'*70}")
print(f"  Sessions where tutorial was dismissed: {sessions_with_game_start} / {total_sessions}")
print(f"  Avg tutorial idle time (TUTORIAL_STARTED → GAME_STARTED): {avg_tutorial_dur:.1f}s  ({avg_tutorial_dur/60:.1f} mins)")
print(f"  (based on {len(tutorial_durations)} sessions with both timestamps)")
print(f"  Avg actual play time (GAME_STARTED → last event): {avg_game_dur:.1f}s  ({avg_game_dur/60:.1f} mins)")
print(f"  (based on {len(game_durations)} sessions with gameplay after tutorial dismiss)")
print(f"  Avg total session time (TUTORIAL_STARTED → last event): {avg_total_dur:.1f}s  ({avg_total_dur/60:.1f} mins)")
print(f"  Avg time from game start to Sun added: {avg_time_to_sun:.1f}s  ({avg_time_to_sun/60:.1f} mins)")
print(f"  (based on {len(sun_times)}/{total_sessions} sessions where Sun was added)")

print(f"\nCASCADES")
print(f"{'-'*70}")
print(f"  Sessions with at least 1 cascade:    {sessions_with_cascade} / {total_sessions}")
print(f"  Avg time to first cascade:           {avg_time_to_cascade:.1f}s  ({avg_time_to_cascade/60:.1f} mins)")
print(f"  (based on {len(all_cascade_times)} sessions with cascades)")
print(f"  Avg animals lost per cascade:        {avg_cascade_size:.1f}")
print(f"  Cascade count distribution:")
for count in sorted(cascade_count_dist):
    label = f"{count} cascade{'s' if count != 1 else ''}"
    print(f"    {label:<15} {cascade_count_dist[count]} session(s)")

print(f"\nFIRST ANIMAL ADDED AFTER SUN")
print(f"{'-'*70}")
print(f"  Most common first animal after Sun:")
for animal, count in first_after_sun_counter.most_common():
    print(f"    {count:>3}x  {animal}")

print(f"\nANIMAL ADD FREQUENCY (all sessions)")
print(f"{'-'*70}")
for animal, count in overall_add_counter.most_common():
    print(f"  {count:>4}x  {animal}")

print(f"\nANIMAL DELETE FREQUENCY (DELETED + CASCADE, all sessions)")
print(f"{'-'*70}")
for animal, count in overall_delete_counter.most_common():
    print(f"  {count:>4}x  {animal}")

most_deleted_overall = overall_delete_counter.most_common(1)
if most_deleted_overall:
    print(f"\n  Most deleted overall: {most_deleted_overall[0][0]} ({most_deleted_overall[0][1]}x)")

# ── Save to Excel ──────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="1A4A6B")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws):
        for cell in ws[ws.max_row]:
            cell.font   = header_font
            cell.fill   = header_fill
            cell.border = border

    def set_col_widths(ws, widths: dict):
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    # ── Overall sheet ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Overall"

    ws1.append(["Metric", "Value"]); style_header_row(ws1)
    ws1.append(["Total Sessions", total_sessions])
    ws1.append(["Sessions Where Tutorial Was Dismissed", sessions_with_game_start])
    ws1.append(["Avg Tutorial Idle Time (s)", round(avg_tutorial_dur, 1)])
    ws1.append(["Avg Tutorial Idle Time (min)", round(avg_tutorial_dur / 60, 2)])
    ws1.append(["Avg Actual Play Time (s)", round(avg_game_dur, 1)])
    ws1.append(["Avg Actual Play Time (min)", round(avg_game_dur / 60, 2)])
    ws1.append(["Avg Total Session Time (s)", round(avg_total_dur, 1)])
    ws1.append(["Avg Total Session Time (min)", round(avg_total_dur / 60, 2)])
    ws1.append(["Avg Time from Game Start to Sun (s)", round(avg_time_to_sun, 1)])
    ws1.append(["Avg Time from Game Start to Sun (min)", round(avg_time_to_sun / 60, 2)])
    ws1.append(["Sessions With Sun Added", len(sun_times)])
    ws1.append(["Sessions With >= 1 Cascade", sessions_with_cascade])
    ws1.append(["Avg Time to First Cascade (s)", round(avg_time_to_cascade, 1)])
    ws1.append(["Avg Animals Lost per Cascade", round(avg_cascade_size, 2)])
    ws1.append([])

    ws1.append(["Device", "Sessions"]); style_header_row(ws1)
    for device, count in device_counter.most_common():
        ws1.append([device, count])
    ws1.append([])

    ws1.append(["Cascade Count per Session", "# Sessions"]); style_header_row(ws1)
    for count in sorted(cascade_count_dist):
        ws1.append([f"{count} cascade(s)", cascade_count_dist[count]])
    ws1.append([])

    ws1.append(["First Animal After Sun", "# Sessions"]); style_header_row(ws1)
    for animal, count in first_after_sun_counter.most_common():
        ws1.append([animal, count])
    ws1.append([])

    ws1.append(["Animal", "Times Added (all sessions)"]); style_header_row(ws1)
    for animal, count in overall_add_counter.most_common():
        ws1.append([animal, count])
    ws1.append([])

    ws1.append(["Animal", "Times Deleted (DELETED + CASCADE, all sessions)"]); style_header_row(ws1)
    for animal, count in overall_delete_counter.most_common():
        ws1.append([animal, count])

    set_col_widths(ws1, {"A": 45, "B": 20})

    # ── Per Session sheet ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Per Session")
    ws2.append([
        "Session", "Device", "IP",
        "Tutorial Started (EDT)", "Game Started (EDT)", "Time Out (EDT)",
        "Tutorial Idle (s)", "Play Time (s)", "Total Time (s)",
        "Time to Sun (s)", "First Animal After Sun",
        "# Cascades", "Longest Cascade (animals)", "Time to First Cascade (s)",
        "Event Type", "Animal", "Cascade #"
    ])
    style_header_row(ws2)

    for s in session_stats:
        i               = s["session_num"]
        device          = s["device"]
        ip              = s["ip"]
        tut_str         = s["tutorial_start"].strftime("%H:%M:%S")
        game_str        = s["game_start"].strftime("%H:%M:%S") if s["game_start"] else "N/A"
        time_out_str    = s["time_out"].strftime("%H:%M:%S")
        tut_dur         = round(s["tutorial_duration"], 1) if s["tutorial_duration"] is not None else "N/A"
        game_dur        = round(s["game_duration"], 1) if s["game_duration"] is not None else "N/A"
        total_dur       = round(s["total_duration"], 1)
        time_to_sun     = round(s["time_to_sun"], 1) if s["time_to_sun"] is not None else "N/A"
        fas             = s["first_after_sun"] or "N/A"
        n_cascades      = len(s["cascade_groups"])
        longest_c       = max(s["cascade_sizes"]) if s["cascade_sizes"] else 0
        ttc             = round(s["time_to_cascade"], 1) if s["time_to_cascade"] is not None else "N/A"

        # Map each cascade event to its cascade number
        cascade_num_map = {}
        for ci, group in enumerate(s["cascade_groups"], 1):
            for ev in group:
                cascade_num_map[id(ev)] = ci

        events = [r for r in s["all_events"] if r["action"] not in SESSION_START_ACTIONS and r["action"] != "GAME_STARTED"]
        first_event = True
        for ev in events:
            cascade_label = cascade_num_map.get(id(ev), "")
            if first_event:
                ws2.append([
                    i, device, ip, tut_str, game_str, time_out_str,
                    tut_dur, game_dur, total_dur,
                    time_to_sun, fas, n_cascades, longest_c, ttc,
                    ev["action"], ev["animal"], cascade_label
                ])
                first_event = False
            else:
                ws2.append([
                    "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                    ev["action"], ev["animal"], cascade_label
                ])

    set_col_widths(ws2, {
        "A": 9, "B": 10, "C": 16, "D": 22, "E": 20, "F": 18,
        "G": 16, "H": 14, "I": 14,
        "J": 16, "K": 28, "L": 12, "M": 22, "N": 22,
        "O": 18, "P": 30, "Q": 10
    })

    out_path = os.path.join(output_dir, "who_eats_whom_analysis.xlsx")
    try:
        wb.save(out_path)
        print(f"\n\nSaved Excel summary → {out_path}")
    except PermissionError:
        # File is open in Excel — save with timestamp to avoid collision
        ts = datetime.now().strftime("%H%M%S")
        out_path = os.path.join(output_dir, f"who_eats_whom_analysis_{ts}.xlsx")
        wb.save(out_path)
        print(f"\n\nNote: who_eats_whom_analysis.xlsx is open — saved to {out_path}")

except ImportError:
    print("\nNote: openpyxl not installed, skipping Excel export.")
    print("Install with: pip install openpyxl")
