# ─────────────────────────────────────────────────────────────────────────────
# analyze_lizard_game.py
#
# Purpose:
#   Analyzes the cleaned pipe-delimited log file for the "Feed the Lizard"
#   game. Produces per-session and overall stats.
#
#   Session splitting strategy:
#     1. Split rows by IP
#     2. Within each IP, split further by device (browser type)
#     3. Within each IP+device stream, split by STARTED events
#     4. All sessions sorted by start time at the end
#
#   Session duration is measured from first drag to last drag to avoid
#   counting idle time before the first animal is touched.
#   All timestamps are converted from UTC to EDT (UTC-4).
#
# Usage:
#   python3 analyze_lizard_game.py <cleaned_game_file>
#   Example: python3 analyze_lizard_game.py cleaned/feed_the_lizard.csv
#
# Output:
#   Prints analysis to console and saves a summary Excel file in cleaned/
#
# ─────────────────────────────────────────────────────────────────────────────
# ASSUMPTIONS & ANALYSIS DECISIONS
# ─────────────────────────────────────────────────────────────────────────────
#
# 1. SESSION IDENTIFICATION
#    A new session is defined by a "SESSION | STARTED" log event. This is
#    triggered every time the game page is loaded or refreshed. Sessions with
#    no drag events following them are retained but noted as having no activity.

# 2. DEVICE SEPARATION (shared WiFi)
#    The event was hosted on a shared WiFi network, meaning multiple devices
#    could share the same IP address. Logs from different devices were therefore
#    interleaved under the same IP in some sessions.
#    To handle this, sessions are first split by IP and then further split by
#    device type extracted from the browser user-agent string (Windows / Mac /
#    iPad / iPhone / Android). Each unique IP+device combination is treated as
#    a separate independent stream before session splitting is applied.

# 3. SESSION DURATION CALCULATION
#    The raw session duration (from STARTED event to last log event) was found
#    to be misleading. In several cases, kids started a session but did not
#    interact with the game immediately, inflating play time significantly
#    (e.g. a 1000s session where actual dragging only happened in the last
#    2 minutes). To correct for this, session duration is calculated as the
#    time from the FIRST drag event to the LAST drag event within the session,
#    excluding any idle time before the game was actually played.

# 4. REPEAT INCORRECTS
#    An animal is flagged as a "repeat incorrect" if a player dragged the same
#    wrong animal to the lizard more than once within a single session. This is
#    used as a signal of confusion or curiosity about that animal.

# 5. CORRECT vs INCORRECT CLASSIFICATION
#    State 1 in the log = correct prey (lizard eats it).
#    State 2 in the log = incorrect prey (lizard rejects it).
#    This is set at the time of logging in the game code and is authoritative.

# 6. TIMEZONE
#    All timestamps in the raw logs are in UTC (from Loki/Grafana).
#    All times in this analysis are converted to EDT (UTC-4), which was the
#    local time at the Marbles Kids Museum event on April 11, 2026 in Raleigh NC.

# 7. TRUNCATED LOG FILE
#    The raw log export from Grafana was limited to 1000 rows. Some browser
#    strings at the end of log lines are truncated at "KHTML" due to the
#    export limit. This does not affect any analysis fields (timestamp, game,
#    ip, animal, action, state) since browser is the last field and is not
#    used for any calculations beyond device type detection, which is reliably
#    determined from the beginning of the browser string.
# ─────────────────────────────────────────────────────────────────────────────

import csv
import os
import sys
from datetime import datetime, timezone, timedelta
from collections import defaultdict, Counter

if len(sys.argv) != 2:
    print("Usage: python3 analyze_lizard_game.py <cleaned_game_file>")
    sys.exit(1)

input_file = sys.argv[1]
output_dir = os.path.dirname(input_file)

EDT = timezone(timedelta(hours=-4))  # EDT = UTC-4 (April)

def get_device(browser: str) -> str:
    if "iPad" in browser:      return "iPad"
    if "iPhone" in browser:    return "iPhone"
    if "Android" in browser:   return "Android"
    if "Windows" in browser:   return "Windows"
    if "Macintosh" in browser: return "Mac"
    return "Unknown"

# ── Load data ─────────────────────────────────────────────────────────────────
rows = []
with open(input_file, newline='', encoding='utf-8') as f:
    reader = csv.DictReader(f, delimiter="|")
    for row in reader:
        row = {k.strip(): v.strip() for k, v in row.items()}
        row["timestamp"] = datetime.fromisoformat(row["timestamp"].replace("Z", "+00:00")).astimezone(EDT)
        row["device"] = get_device(row["browser"])
        rows.append(row)

rows.sort(key=lambda r: r["timestamp"])

# ── Step 1: Split by IP, then by device ───────────────────────────────────────
streams: dict = defaultdict(list)
for row in rows:
    key = (row["ip"], row["device"])
    streams[key].append(row)

# ── Step 2: Within each stream, split by STARTED events ───────────────────────
all_sessions = []
for (ip, device), stream_rows in streams.items():
    current = []
    for row in stream_rows:
        if row["action"] == "STARTED":
            if current:
                all_sessions.append(current)
            current = [row]
        else:
            current.append(row)
    if current:
        all_sessions.append(current)

# ── Step 3: Sort all sessions by start time ───────────────────────────────────
all_sessions.sort(key=lambda s: s[0]["timestamp"])

print(f"{'='*65}")
print(f"  FEED THE LIZARD — LOG ANALYSIS  (all times in EDT)")
print(f"{'='*65}\n")

# ── OVERALL STATS ─────────────────────────────────────────────────────────────
print(f"OVERALL")
print(f"{'-'*65}")

total_sessions = len(all_sessions)
device_counter = Counter(s[0]["device"] for s in all_sessions)
print(f"  Number of sessions:        {total_sessions}")
print(f"  Sessions by device:")
for device, count in device_counter.most_common():
    print(f"    {device:<12} {count} sessions")

correct_counter   = Counter()
incorrect_counter = Counter()
session_durations = []

for session in all_sessions:
    drags = [r for r in session if r["action"] == "DRAGGED"]
    for d in drags:
        if d["state"] == "1":   correct_counter[d["animal"]] += 1
        elif d["state"] == "2": incorrect_counter[d["animal"]] += 1
    # Duration = first drag to last drag (excludes idle time before first touch)
    if drags:
        session_durations.append((drags[-1]["timestamp"] - drags[0]["timestamp"]).total_seconds())

avg_duration = sum(session_durations) / len(session_durations) if session_durations else 0
print(f"\n  Avg time spent on game:    {avg_duration:.1f}s  ({avg_duration/60:.1f} mins)")

print(f"\n  Most correctly identified animals:")
for animal, count in correct_counter.most_common(5):
    print(f"    {count:>4}x  {animal}")

print(f"\n  Most incorrectly dragged animals:")
for animal, count in incorrect_counter.most_common(5):
    print(f"    {count:>4}x  {animal}")

# ── PER SESSION STATS ─────────────────────────────────────────────────────────
print(f"\n\nPER SESSION")
print(f"{'-'*65}")

for i, session in enumerate(all_sessions, 1):
    drag_times = [r["timestamp"] for r in session if r["action"] == "DRAGGED"]
    session_start = session[0]["timestamp"]
    first_drag    = drag_times[0] if drag_times else session_start
    last_drag     = drag_times[-1] if drag_times else session_start
    duration      = (last_drag - first_drag).total_seconds()
    drags         = [r for r in session if r["action"] == "DRAGGED"]
    ip            = session[0]["ip"]
    device        = session[0]["device"]

    print(f"\n  Session {i}  |  {device}  |  IP: {ip}  |  {first_drag.strftime('%H:%M:%S EDT')} -> {last_drag.strftime('%H:%M:%S EDT')}  |  {duration:.0f}s")

    if not drags:
        print(f"    No drags recorded")
        continue

    print(f"    Order dragged:")
    for d in drags:
        result = "CORRECT" if d["state"] == "1" else "WRONG"
        print(f"      {d['timestamp'].strftime('%H:%M:%S')}  {d['animal']:<30} {result}")

    wrong_counts = Counter(d["animal"] for d in drags if d["state"] == "2")
    repeats = {a: c for a, c in wrong_counts.items() if c > 1}
    if repeats:
        print(f"    Repeat incorrects:")
        for animal, count in sorted(repeats.items(), key=lambda x: -x[1]):
            print(f"      {animal}: dragged incorrectly {count}x")
    else:
        print(f"    No repeat incorrects")

# ── Save to Excel ──────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="2D5016")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws):
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

    # Overall sheet
    ws1 = wb.active
    ws1.title = "Overall"
    ws1.append(["Metric", "Value"]); style_header_row(ws1)
    ws1.append(["Number of Sessions", total_sessions])
    ws1.append(["Avg Time Spent (seconds)", round(avg_duration, 1)])
    ws1.append(["Avg Time Spent (minutes)", round(avg_duration/60, 2)])
    ws1.append([])
    ws1.append(["Device", "Sessions"]); style_header_row(ws1)
    for device, count in device_counter.most_common():
        ws1.append([device, count])
    ws1.append([])
    ws1.append(["Most Correct Animal", "Count"]); style_header_row(ws1)
    for animal, count in correct_counter.most_common():
        ws1.append([animal, count])
    ws1.append([])
    ws1.append(["Most Incorrect Animal", "Count"]); style_header_row(ws1)
    for animal, count in incorrect_counter.most_common():
        ws1.append([animal, count])
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 12

    # Per session sheet
    ws2 = wb.create_sheet("Per Session")
    ws2.append(["Session", "Device", "IP", "First Drag (EDT)", "Last Drag (EDT)", "Duration (s)", "Animal", "Result"])
    style_header_row(ws2)
    for i, session in enumerate(all_sessions, 1):
        drag_times = [r["timestamp"] for r in session if r["action"] == "DRAGGED"]
        first_drag = drag_times[0] if drag_times else session[0]["timestamp"]
        last_drag  = drag_times[-1] if drag_times else session[0]["timestamp"]
        duration   = (last_drag - first_drag).total_seconds()
        ip         = session[0]["ip"]
        device     = session[0]["device"]
        drags      = [r for r in session if r["action"] == "DRAGGED"]
        if not drags:
            ws2.append([i, device, ip, first_drag.strftime("%H:%M:%S"), last_drag.strftime("%H:%M:%S"), round(duration), "—", "—"])
        for d in drags:
            result = "Correct" if d["state"] == "1" else "Wrong"
            ws2.append([i, device, ip, first_drag.strftime("%H:%M:%S"), last_drag.strftime("%H:%M:%S"),
                        round(duration), d["animal"], result])
    for col in ["A","B","C","D","E","F","G","H"]:
        ws2.column_dimensions[col].width = 20

    out_path = os.path.join(output_dir, "lizard_analysis.xlsx")
    wb.save(out_path)
    print(f"\n\nSaved Excel summary to: {out_path}")

except ImportError:
    print("\nNote: openpyxl not installed, skipping Excel export.")